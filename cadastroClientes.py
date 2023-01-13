import PySimpleGUI as sg
import requests as rs
from openpyxl import workbook, load_workbook

# Layout
sg.change_look_and_feel('DarkAmber')

layout = [
        [sg.Text('Nome', size=(4, 0)), sg.Input(key='nome', size=(20, 0)), sg.Text('Data de Nascimento', size=(15,0)), sg.Input(key='dataNascimento', size=(10 ,0))],
        [sg.Text('CPF',size=(4, 0)), sg.Input(key='cpf', size=(15, 0)), sg.Text('RG', size=(2, 0)), sg.Input(key='rg', size=(15,0))],
        [sg.Text('CEP', size=(4, 0)), sg.Input(key='cep', size=(10, 0)), sg.Button('Consultar')],
        [sg.Text('Logradouro', size=(8, 0)), sg.Input(key='logradouro', size=(25, 0)), sg.Text('Número', size=(6, 0)), sg.Input(key='número', size=(6, 0))],
        [sg.Text('Complemento', size=(10, 0)), sg.Input(key='complemento', size=(8, 0)), sg.Text('Bairro', size=(5, 0)), sg.Input(key='bairro', size=(15, 0))],
[       sg.Text('Localidade', size=(10, 0)), sg.Input(key='localidade', size=(10, 0)), sg.Text('Estado', size=(5, 0)), sg.Input(key='uf', size=(3, 0))],
        [sg.Button('Cadastrar'), sg.Button('Sair'), sg.Text('', key='mensagem', size=(33, 0))]
]

# Janela
janela1 = sg.Window('Clientes', layout)

#Leitura de dados
def Iniciar():
        while True:
                eventos, valores = janela1.read()

                if eventos == sg.WINDOW_CLOSED or eventos == 'Sair':
                        break

                if eventos == 'Cadastrar':
                        if validacoes(valores['nome'], valores['dataNascimento'], valores['rg'], valores['cpf']):
                                cadastrarCliente(valores['nome'], valores['dataNascimento'], valores['rg'], valores['cpf'])
                                limparFormulário()
                if eventos == 'Consultar':
                        consultarCEP(valores['cep'])

def limparFormulário():
        janela1['nome'].update('')
        janela1['dataNascimento'].update('')
        janela1['rg'].update('')
        janela1['cpf'].update('')

def cadastrarCliente(nome, dataNascimento, rg, cpf):
        wb = load_workbook('Clientes.xlsx')
        ws = wb.active
        for cell in range(2, 1048576):
                if ws['A' + str(cell)].value == None:
                        ws['A' + str(cell)].value = nome
                        ws['B' + str(cell)].value = dataNascimento
                        ws['C' + str(cell)].value = rg
                        ws['D' + str(cell)].value = cpf
                        janela1['mensagem'].update('Cliente Cadastrado com Sucesso !', text_color='green')
                        wb.save('Clientes.xlsx')
                        break

def validacoes(nome, dataNascimento, rg, cpf):
        if str(nome).strip() == '':
                janela1['mensagem'].update('Digite o nome do Cliente !', text_color='red')
                return False
        if str(dataNascimento).strip() == '':
                janela1['mensagem'].update('Digite a Data de Nascimento !', text_color='red')
                return False
        if str(cpf).strip == '':
                janela1['mensagem'].update('Digite o CPF !', text_color='red')
                return False
        if str(rg).strip() == '':
                janela1['mensagem'].update('Digite o RG !', text_color='red')
                return False

        return True

def consultarCEP(cep):
        if len(cep) != 8:
                print('Quantidade de dígitos inválida!')
                exit()
        request = rs.get('https://viacep.com.br/ws/{}/json/'.format(cep))

        address_data = request.json()

        if 'erro' not in address_data:
                janela1['logradouro'].update(address_data['logradouro'])
                janela1['bairro'].update(address_data['bairro'])
                janela1['localidade'].update(address_data['localidade'])
                janela1['uf'].update(address_data['uf'])


